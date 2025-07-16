import pendulum
import requests
from bs4 import BeautifulSoup
import os
import logging
import subprocess

from airflow.decorators import dag, task


from neo4j import GraphDatabase
from langchain_openai import OpenAIEmbeddings
from dotenv import load_dotenv

from pypdf import PdfReader
from openpyxl import load_workbook
from docx import Document

airflow_home = os.environ.get('AIRFLOW_HOME', '~/airflow')
dotenv_path = os.path.join(os.path.expanduser(airflow_home), '.env')
load_dotenv(dotenv_path=dotenv_path)

# 환경 변수 로드
NEO4J_URI = os.getenv("NEO4J_URI")
NEO4J_USERNAME = os.getenv("NEO4J_USERNAME")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# 임베딩 모델 전역 초기화
embeddings = OpenAIEmbeddings(model="text-embedding-3-small", openai_api_key=OPENAI_API_KEY)


def get_neo4j_driver():
    """Neo4j 드라이버 인스턴스를 반환합니다."""
    return GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USERNAME, NEO4J_PASSWORD))


def extract_text_from_file(file_path):
    """
    파일 경로를 받아 확장자에 따라 텍스트를 추출하는 함수.
    HWP 파일은 'hwp5-to-text' 커맨드라인 도구를 사용하여 처리합니다.
    """
    _, extension = os.path.splitext(file_path.lower())
    text = ""
    try:
        if extension == '.hwp':
            result = subprocess.run(
                ['hwp5-to-text', file_path],  # 실행할 명령어: hwp5-to-text "hwp파일명"
                capture_output=True,  # 표준 출력(stdout)과 표준 에러(stderr)를 캡처
                text=True,  # 결과를 문자열(text)로 디코딩
                check=True,  # 명령어 실패 시 예외 발생
                encoding='utf-8'  # UTF-8로 인코딩
            )
            # 캡처된 표준 출력(stdout)을 text 변수에 저장
            text = result.stdout
        elif extension == '.pdf':
            reader = PdfReader(file_path)
            for page in reader.pages:
                text += page.extract_text() or ""
        elif extension == '.docx':
            doc = Document(file_path)
            for para in doc.paragraphs:
                text += para.text + '\n'
        elif extension == '.xlsx':
            workbook = load_workbook(filename=file_path, data_only=True)
            for sheet in workbook:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text += str(cell.value) + ' '
                    text += '\n'
        logging.info(f"Successfully extracted text from {file_path}")
    except FileNotFoundError:
        logging.error(f"Command 'hwp5-to-text' not found. Please install it using 'pip install hwp5-to-text'.")
    except subprocess.CalledProcessError as e:
        logging.error(f"hwp5-to-text failed for {file_path}: {e.stderr}")
    except Exception as e:
        logging.error(f"Error extracting text from {file_path}: {e}")
    return text


@dag(
    dag_id='knut_announcement_pipeline',
    start_date=pendulum.datetime(2024, 5, 20, tz="Asia/Seoul"),
    schedule='@daily',
    catchup=False,
    tags=['knut', 'chatbot', 'neo4j'],
)
def school_announcement_pipeline_neo4j():
    @task
    def setup_database_constraints():
        """Neo4j에 제약조건과 벡터 인덱스를 생성합니다."""
        logging.info("Setting up Neo4j constraints and vector index...")
        driver = get_neo4j_driver()
        with driver.session() as session:
            session.run("CREATE CONSTRAINT announcement_url IF NOT EXISTS FOR (a:Announcement) REQUIRE a.url IS UNIQUE")

            try:
                session.run("""
                    CREATE VECTOR INDEX announcement_embeddings IF NOT EXISTS
                    FOR (a:Announcement) ON (a.embedding)
                    OPTIONS { indexConfig: {
                        `vector.dimensions`: 1536,
                        `vector.similarity_function`: 'cosine'
                    }}
                """)
                logging.info("Vector index 'announcement_embeddings' is ready.")
            except Exception as e:
                logging.warning(f"Vector index may already exist: {e}")
        driver.close()

    @task
    def scrape_announcements() -> list[dict]:
        """
        국립한국교통대학교 일반소식 게시판을 스크래핑하여
        올바른 상세 페이지 URL을 생성하는 태스크.
        """
        BASE_URL = "https://www.ut.ac.kr"
        BOARD_URL = f"{BASE_URL}/cop/bbs/BBSMSTR_000000000059/selectBoardList.do"

        logging.info(f"Scraping started for: {BOARD_URL}")

        try:
            response = requests.get(BOARD_URL)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            announcements = []

            table_body = soup.select_one('table.basic_table > tbody')
            if not table_body:
                logging.warning("Could not find the announcement table body.")
                return []

            for row in table_body.find_all('tr'):
                # 고정 공지는 a 태그, 일반 공지는 form 태그로 되어 있어 분기 처리

                # 1. 고정 공지 처리 (<a> 태그)
                notice_link = row.select_one('td.left div.list_subject a')
                if notice_link:
                    title = notice_link.text.strip()
                    link = notice_link.get('href', '')

                    if title and link:
                        full_link = requests.compat.urljoin(BASE_URL, link)
                        # 중복 추가 방지
                        if not any(ann['url'] == full_link for ann in announcements):
                            announcements.append({'title': title, 'url': full_link})
                    continue  # 고정 공지 처리 후 다음 행으로

                # 2. 일반 공지 처리 (<form> 태그)
                form_tag = row.select_one('td.left div.list_subject form')
                if form_tag:
                    action_url = form_tag.get('action')

                    # form 내부의 모든 input 태그를 딕셔너리로 만듦
                    params = {
                        inp.get('name'): inp.get('value')
                        for inp in form_tag.find_all('input') if inp.get('name')
                    }

                    title = params.get('nttSj') or form_tag.select_one('input[type="submit"]').get('value', '').strip()

                    if title and action_url and 'nttId' in params:
                        # URL 쿼리 스트링으로 변환 (예: nttId=1234&bbsId=...)
                        query_string = requests.compat.urlencode(params)
                        # 완전한 URL 조립
                        full_link = f"{BASE_URL}{action_url}?{query_string}"

                        if not any(ann['url'] == full_link for ann in announcements):
                            announcements.append({'title': title, 'url': full_link})

            logging.info(f"Scraped {len(announcements)} announcements.")
            # 디버깅을 위해 추출된 URL 중 하나를 로그로 출력
            if announcements:
                logging.info(f"Example URL: {announcements[0]['url']}")

            return announcements

        except requests.exceptions.RequestException as e:
            logging.error(f"Failed to retrieve the webpage {BOARD_URL}. Error: {e}")
            return []
        except Exception as e:
            logging.error(f"An error occurred during scraping: {e}", exc_info=True)
            return []

    @task
    def process_and_store_in_neo4j(announcement: dict):
        """
        개별 공지사항 페이지를 파싱하고, 본문과 첨부파일 내용을 추출하여 Neo4j에 저장하는 태스크.
        파일명 추출 로직을 수정하여 정확한 이름만 가져옵니다.
        """
        try:
            url = announcement['url']
            title = announcement['title']
            logging.info(f"Processing: {title} ({url})")

            with requests.Session() as s:
                response = s.get(url)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, 'html.parser')

                # 본문 내용 추출
                content_div = soup.select_one('div.bbs-view-content')
                content = content_div.get_text(separator='\n', strip=True) if content_div else ""

                # 첨부파일 처리
                file_content = ""
                temp_dir = "/tmp/school_files"
                os.makedirs(temp_dir, exist_ok=True)

                attachment_div = soup.select_one('div.bbs_detail_file')
                if attachment_div:
                    for file_link_tag in attachment_div.find_all('a'):
                        file_params_div = file_link_tag.find_next_sibling('div', style="display: none")
                        if not file_params_div:
                            continue

                        file_params = file_params_div.get_text(strip=True)

                        # --- ★★★ 여기가 수정된 부분입니다 ★★★ ---
                        # <a> 태그의 자식 노드 중 첫 번째 요소(텍스트)만 가져와서 공백을 제거합니다.
                        file_name = file_link_tag.contents[0].strip()
                        # ----------------------------------------

                        base_download_url = "https://www.ut.ac.kr/cmm/fms/FileDown.do"
                        full_file_url = f"{base_download_url}{file_params}"

                        logging.info(f"Downloading file: '{file_name}' from {full_file_url}")

                        try:
                            file_response = s.get(full_file_url)
                            file_response.raise_for_status()

                            temp_file_path = os.path.join(temp_dir, file_name)
                            with open(temp_file_path, 'wb') as f:
                                f.write(file_response.content)

                            file_text = extract_text_from_file(temp_file_path)
                            # 깔끔하게 추출된 파일 이름을 사용합니다.
                            file_content += f"\n\n--- 첨부파일: {file_name} ---\n{file_text}"

                            os.remove(temp_file_path)
                        except Exception as file_e:
                            logging.error(f"Failed to process file {file_name} from {url}: {file_e}")

                # 임베딩 및 DB 저장
                full_text = f"제목: {title}\n\n본문:\n{content}\n\n첨부파일 내용:\n{file_content}"
                if len(full_text) > 8000:
                    full_text = full_text[:8000]

                embedding_vector = embeddings.embed_query(full_text)

                driver = get_neo4j_driver()
                with driver.session() as db_session:
                    db_session.run("""
                        MERGE (a:Announcement {url: $url})
                        SET a.title = $title,
                            a.content = $content,
                            a.file_content = $file_content,
                            a.full_text = $full_text,
                            a.embedding = $embedding,
                            a.createdAt = datetime()
                    """, {
                        "url": url, "title": title, "content": content,
                        "file_content": file_content, "full_text": full_text, "embedding": embedding_vector
                    })
                driver.close()
                logging.info(f"Successfully stored in Neo4j: {title}")

        except Exception as e:
            logging.error(f"Failed to process URL {announcement.get('url', 'N/A')}: {e}", exc_info=True)

    # 태스크 실행 순서 정의
    setup_task = setup_database_constraints()
    scraped_list = scrape_announcements()
    setup_task >> scraped_list
    process_and_store_in_neo4j.expand(announcement=scraped_list)


school_announcement_pipeline_neo4j()